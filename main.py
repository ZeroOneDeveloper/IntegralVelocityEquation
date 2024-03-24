import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from typing import List, Tuple, Any
from sklearn.linear_model import LinearRegression

try:
    filename = input("Please enter the file name to test.\n(ex. sample.xlsx) : ")
    workbook = openpyxl.load_workbook(f"./datasets/{filename}")
except FileNotFoundError:
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        print("File not found. Please check the file name and try again.")
        exit()
try:
    max_reaction_order = int(
        input("Please enter the maximum reaction order to test.\n(ex. 10) : ")
    )
    if max_reaction_order < 0:
        raise ValueError
except ValueError:
    print("Invalid input. Please enter a natural number.")
    exit()

sheet = workbook.active

data: List[Any] = [[cell[0].value, cell[1].value] for cell in sheet.iter_rows()]


def reaction(
    n: int, inputData: List[List[Any]]
) -> Tuple[np.array, np.array, np.ndarray, float]:
    t = np.array(list(map(lambda x: x[0], inputData)))
    if n == 0:
        A = np.array(list(map(lambda x: x[1], inputData)))
    elif n == 1:
        A = np.log(np.array(list(map(lambda x: x[1], inputData))))
    elif n >= 2:
        A = np.array(list(map(lambda x: 1 / x[1] ** (n - 1), inputData)))

    model = LinearRegression()
    model.fit(X=t.reshape(-1, 1), y=A)

    pred = model.predict(t.reshape(-1, 1))

    r2 = model.score(t.reshape(-1, 1), A)

    return t, A, pred, r2


def test():
    r2_scores = []
    r2_results = openpyxl.Workbook()
    r2_sheet = r2_results.active
    for n in range(0, max_reaction_order + 1):
        t, A, pred, r2 = reaction(n, data)

        plt.scatter(t, A)
        plt.title(f"Reaction Order {n}")
        plt.xlabel("Time")
        plt.ylabel("Concentration")
        # plt figtext on right bottom
        plt.figtext(0.99, 0.01, f"R² Score: {r2}", horizontalalignment="right")
        plt.plot(t, pred, color="red")
        plt.savefig(f"./results/{n}.png")
        plt.close()

        r2_scores.append(r2)
        r2_sheet.append([n, r2])
    r2_results.save(f"{filename[:-5]}_r2.xlsx")
    print(
        f"The best reaction order is {r2_scores.index(max(r2_scores))} with R² score of {max(r2_scores)}\n"
        "Please check the results.xlsx for the graphs and the R^2 scores."
    )


test()
